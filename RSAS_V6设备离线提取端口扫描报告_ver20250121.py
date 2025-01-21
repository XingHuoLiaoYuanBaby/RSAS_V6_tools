# coding=utf-8
# 导入所需的标准库
from math import floor  # 用于进度计算时向下取整
from os import path, getcwd, listdir, system  # 文件和系统操作相关函数
from pathlib import Path  # 路径处理
from re import match  # 正则表达式匹配
from time import strftime, localtime  # 时间处理
from zipfile import ZipFile  # ZIP文件处理

# 导入第三方库
from colorama import init, Fore  # 控制台颜色输出
from openpyxl import Workbook  # Excel写入
from xlrd import open_workbook  # Excel读取


# initiate font color
init(autoreset=True)


def readZipData(path_, filename):
    """
    从ZIP文件中读取RSAS扫描报告的XLS文件
    Args:
        path_: ZIP文件所在路径
        filename: ZIP文件名
    Returns:
        dic_list: 包含所有主机信息和端口信息的字典列表
    """
    dic_list = []
    with ZipFile(Path(f'{path_}/{filename}'), 'r') as f:
        for name in f.namelist():
            # 匹配形如 x.x.x.x.xls 的文件名（IP地址格式）
            if match(r'(\d+\.){4}xls', name):
                with f.open(name, mode='r') as data:
                    content = open_workbook(file_contents=data.read())
                    dic_list.append(readPortXlsData(data=content))
    return dic_list


#   read data from xls, xlrd column and row begin at 0
def readPortXlsData(filename='', data=''):
    """
    读取并解析RSAS扫描报告中的端口信息
    Args:
        filename: 文件名（可选）
        data: 已打开的workbook对象（可选）
    Returns:
        dic: 包含主机信息和端口信息的字典
    """
    # 如果提供了文件名，则打开文件
    if filename:
        data = open_workbook(filename)
    
    # 获取两个关键sheet页
    host_data = data.sheet_by_name('主机概况')  # 包含主机基本信息的sheet
    port_data = data.sheet_by_name('其它信息')  # 包含端口信息的sheet
    row_count = port_data.nrows  # 获取行数
    
    # 初始化数据存储结构
    dic = {}  # 用于存储所有提取的信息
    port_info_list = []  # 用于存储端口信息的列表
    
    # 提取主机基本信息
    dic[0] = host_data.cell(2, 1).value  # 主机IP地址
    
    # 初始化主机名和操作系统列索引
    numbers = [1, 2, 3, 4]
    h = 0  # 主机名列索引
    s = 0  # 操作系统列索引
    
    # 查找主机名和操作系统所在列
    for i in numbers:
        if host_data.cell(4, i).value == u'主机名':
            h = i
        elif host_data.cell(4, i).value == u'操作系统':
            s = i
    
    # 提取主机名，如果未找到则置空
    if h == 0:
        dic[1] = ' '
    else:
        dic[1] = host_data.cell(5, h).value  # 主机名
    
    # 提取操作系统信息，如果未找到则置空
    if s == 0:
        dic[2] = ' '
    else:
        dic[2] = host_data.cell(5, s).value  # 操作系统
    
    # 提取扫描时间
    dic[3] = host_data.cell(8, 2).value  # 扫描完成时间
    # 如果扫描时间格式不正确，使用开始时间
    if not match('\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', dic[3]):
        dic[3] = host_data.cell(8, 1).value  # 使用扫描开始时间
    
    # 端口信息处理部分
    port_start_row = -1  # 端口信息起始行
    port_info_found = False  # 是否找到端口信息标记
    
    try:
        # 查找端口信息的起始位置
        for i in range(row_count):
            if port_data.cell(i, 0).value == '远程端口信息':
                port_start_row = i + 2  # 跳过标题行（+2：跳过标题"远程端口信息"和列头）
                port_info_found = True
                break
        
        # 如果没有找到端口信息，返回空的端口列表
        if not port_info_found:
            dic[4] = port_info_list
            return dic
            
        # 处理端口信息
        if port_start_row > 0:
            for row in range(port_start_row, row_count):
                row_data = port_data.row_values(row)
                # 遇到空行时结束读取（端口列为空表示数据结束）
                if not row_data[1]:
                    break
                
                # 处理端口信息
                try:
                    port_value = str(row_data[1]).strip()
                    # 处理端口范围情况（例如：80-89）
                    if '-' in port_value:
                        # 解析起始端口和结束端口
                        start_port, end_port = map(int, port_value.split('-'))
                        # 遍历端口范围，为每个端口创建记录
                        for port in range(start_port, end_port + 1):
                            port_info = [
                                row_data[0],  # 传输层协议（TCP/UDP）
                                port,  # 端口号
                                row_data[2],  # 应用层协议
                                row_data[3],  # 服务名称
                                row_data[4] if len(row_data) > 4 else 'unknown'  # 端口状态，如果没有则标记为unknown
                            ]
                            port_info_list.append(port_info)
                    else:
                        # 处理单个端口
                        if port_value and (isinstance(row_data[1], (int, float)) or 
                            (isinstance(port_value, str) and port_value.strip())):
                            port_info = [
                                row_data[0],  # 传输层协议（TCP/UDP）
                                int(float(port_value)) if port_value.replace('.', '').isdigit() else port_value,  # 端口号
                                row_data[2],  # 应用层协议
                                row_data[3],  # 服务名称
                                row_data[4] if len(row_data) > 4 else 'unknown'  # 端口状态
                            ]
                            port_info_list.append(port_info)
                except (ValueError, TypeError) as e:
                    continue  # 跳过无效的端口数据
                    
    except Exception as e:
        # 处理异常情况，打印错误信息并返回已收集的数据
        print(f'\t{Fore.RED}[-]{current_time()}\t处理端口信息时发生错误：{str(e)}')
        dic[4] = port_info_list
        
    # 存储端口信息列表并返回
    dic[4] = port_info_list
    return dic


# save data as .xlsx, openpyxl column and row begin at 1
def save(path_, file, dic_list):
    """
    将处理后的数据保存为xlsx格式
    Args:
        path_: 输出文件路径
        file: 原始ZIP文件名
        dic_list: 包含所有主机和端口信息的字典列表
    """
    # 获取不带扩展名的文件名
    file_name = path.splitext(file)[0]
    
    # 计算总行数（包含所有端口信息和没有端口信息的主机，加上标题行）
    total = sum([len(dic[4]) for dic in dic_list]) + sum([1 for dic in dic_list if len(dic[4]) == 0]) + 1
    
    # 从文件名中提取任务ID
    taskid = int(file_name.split('_')[0])
    
    # 创建新的Excel工作簿
    output_xlsx = Workbook()
    sheet = output_xlsx.active
    
    # 写入表头（新增 ip:port 列）
    sheet.cell(row=1, column=1, value='taskid')
    sheet.cell(row=1, column=2, value='ip')
    sheet.cell(row=1, column=3, value='hostname')
    sheet.cell(row=1, column=4, value='system_type')
    sheet.cell(row=1, column=5, value='scan_time')
    sheet.cell(row=1, column=6, value='port')
    sheet.cell(row=1, column=7, value='protocol')
    sheet.cell(row=1, column=8, value='service')
    sheet.cell(row=1, column=9, value='status')
    sheet.cell(row=1, column=10, value='ip:port')  # 新增列
    
    count = 1  # 当前处理的行数
    feedback = 0  # 进度反馈
    
    # 遍历所有主机数据
    for dic_data in dic_list:
        # 处理没有端口信息的主机
        if not (len(dic_data[4])):
            # 写入主机基本信息
            sheet.cell(row=count+1, column=1, value=taskid)
            sheet.cell(row=count+1, column=2, value=dic_data[0])
            sheet.cell(row=count+1, column=3, value=dic_data[1])
            sheet.cell(row=count+1, column=4, value=dic_data[2])
            sheet.cell(row=count+1, column=5, value=dic_data[3])
            # 端口相关字段填充null
            i = 6
            while i <= 10:  # 修改循环范围以包含新列
                sheet.cell(row=count+1, column=i, value='null')
                i += 1
            count += 1
        else:
            # 处理有端口信息的主机
            for i in range(len(dic_data[4])):
                # 写入主机基本信息
                sheet.cell(row=count + 1, column=1).value = taskid
                sheet.cell(row=count + 1, column=2).value = dic_data[0]
                sheet.cell(row=count + 1, column=3).value = dic_data[1]
                sheet.cell(row=count + 1, column=4).value = dic_data[2]
                sheet.cell(row=count + 1, column=5).value = dic_data[3]
                # 写入端口信息
                sheet.cell(row=count + 1, column=6).value = dic_data[4][i][1]
                sheet.cell(row=count + 1, column=7).value = dic_data[4][i][2]
                sheet.cell(row=count + 1, column=8).value = dic_data[4][i][3]
                sheet.cell(row=count + 1, column=9).value = dic_data[4][i][4]
                # 写入 ip:port 组合
                sheet.cell(row=count + 1, column=10).value = f"{dic_data[0]}:{dic_data[4][i][1]}"
                count += 1
        # 更新进度显示
        feedback = progress(file_name, count, total, feedback, '读取')
    
    # 保存文件
    output = Path(f'{path_}/{file_name}.xlsx')
    print(f'\t{Fore.GREEN}[+]{current_time()}\t正在保存文件，请稍候……')
    output_xlsx.save(filename=output)
    output_xlsx.close()
    print(f'\t{Fore.GREEN}[+]{current_time()}\t文件[{file_name}]保存完毕。')


# current time
def current_time():
    """
    获取当前时间的格式化字符串
    Returns:
        str: 格式化的时间字符串 (YYYY-MM-DD HH:MM:SS)
    """
    return strftime("%Y-%m-%d %H:%M:%S", localtime())


# show progress
def progress(file_name, count, total, feedback, type):
    """
    显示处理进度
    Args:
        file_name: 当前处理的文件名
        count: 当前处理的数量
        total: 总数量
        feedback: 上一次反馈的进度值
        type: 处理类型（如：'读取'）
    Returns:
        int: 当前进度百分比
    """
    pct = floor(count / total * 100)  # 计算百分比并向下取整
    if pct > feedback:  # 只在进度增加时更新显示
        print(f'\r\t{Fore.GREEN}[+]{current_time()}\t文件[{file_name}]\t当前{type}进度[{count} / {total}  {pct:<.0f}%]', end='')
    if pct == 100:  # 处理完成时换行
        print()
    return pct


# main program
def main():
    """
    主程序入口
    处理pending目录下的所有符合命名规则的ZIP文件
    """
    files = []
    # 获取pending目录路径
    path_ = Path(f'{getcwd()}/pending')
    
    # 查找符合命名规则的ZIP文件
    for filename in listdir(path_):
        # 匹配文件名格式：数字_名称_年_月_日_xls.zip 或 数字_名称_年_月_日_excel.zip
        if match(r'\d+_\S+_\d{4}_\d{2}_\d{2}_xls\.zip', filename) or match(r'\d+_\S+_\d{4}_\d{2}_\d{2}_excel\.zip', filename):
            files.append(filename)
    
    # 处理每个找到的文件
    for file in files:
        print(f'{Fore.GREEN}[*]{current_time()}\t正在处理第 {files.index(file) + 1}/{len(files)} 个文件。')
        dic_list = readZipData(path_, file)
        save(path_, file, dic_list)
    
    print(f'{Fore.GREEN}[*]{current_time()}\t所有数据已处理完毕。')
    system('pause')  # 等待用户按键后退出


if __name__ == "__main__":
    main()
